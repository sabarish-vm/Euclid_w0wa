import numpy as np
import openpyxl
from glob import glob,os
from itertools import combinations as cb

os.chdir(os.path.dirname(os.path.realpath(__file__)))

def updater(sheet,initialcell,value,i,j) :
    row = sheet[initialcell].row + i
    col = sheet[initialcell].column + j
    sheet.cell(column=col,row=row).value = value
  
    row = sheet[initialcell].row + j
    col = sheet[initialcell].column + i
    sheet.cell(column=col,row=row).value = value

def write(files_list,sheet,code_one,code_two):
    for file in files_list :
        filename = os.path.basename(file)
        if code_one in filename :
            i = code_order[code_one]
            if code_two in filename :
                j = code_order[code_two]
                for typ in name_to_cell.keys() :
                    if typ in filename :
                        contents = np.loadtxt(file)
                        with open(file,'r') as f :
                            maxval=np.amax(np.abs(contents))
                            updater(sheet,name_to_cell[typ],maxval,i,j)
                    



###############################################

xl = openpyxl.load_workbook('./template.xlsx')
photo_opt = xl['Photometric-Opt']
spec_opt = xl['Spectroscopic-Opt']
photo_pess = xl['Photometric-Pess']
spec_pess = xl['Spectroscopic-Pess']
files = glob('../../plots/*/*/*error*.txt')

###############################################

name_to_cell = {'cosmo_and_nuisance' : 'A3', 'cosmo_fix_nuisance' : 'H3',
'cosmo_marg_nuisance' : 'A13', 'nuisance_fix_cosmo' : 'H13'}
code_order = {'class_ext':1,'class_int':2,'MP':3,'camb_ext' : 4, 'camb_int':5}
all_combinations = list(cb(code_order.keys(),2))

#################################################

photo_opt_files = []
photo_pess_files = []
spec_opt_files = []
spec_pess_files = []
for file in files :
    if 'photometric' in file :
        if 'optimistic' in file :
            photo_opt_files.append(file)
        elif 'pessimistic' in file :
            photo_pess_files.append(file)
    elif 'spectroscopic' in file :
        if 'optimistic' in file :
            spec_opt_files.append(file)
        elif 'pessimistic' in file :
            spec_pess_files.append(file)

###################################################             
for i in all_combinations :
    write(photo_opt_files,photo_opt,i[0],i[1])
    write(photo_pess_files,photo_pess,i[0],i[1])
    write(spec_opt_files,spec_opt,i[0],i[1])
    write(spec_pess_files,spec_pess,i[0],i[1])

####################################################

xl.save('Fishers_comparison.xlsx')
